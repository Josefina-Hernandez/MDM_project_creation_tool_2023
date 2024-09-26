from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException, StaleElementReferenceException, ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options

import time
import datetime

import os

import openpyxl as xl

class Excel_Con():
    def __init__(self, file_name):
        self.wb = xl.load_workbook(filename=file_name, data_only=True)
        self.ws = self.wb.active

        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime("%Y%m%d%H%M%S")
        self.output_name = 'output_' + formatted_time + '.xlsx'


    def check_max_row(self):
        for i in range(10, self.ws.max_row+1):
            if not self.ws.cell(i, 1).value:
                return i-1

        return self.ws.max_row

    def reading_specs(self):
        ID = str(self.ws['F3'].value).strip()
        PW = str(self.ws['F4'].value).strip()
        subsidiary = str(self.ws['F5'].value).strip()
        country = str(self.ws['H5'].value).strip()
        ui_language = str(self.ws['F6'].value).strip()

        return {
            'ID': ID,
            'PW': PW,
            'subsidiary': subsidiary,
            'country': country,
            'ui_language': ui_language
        }

    def excel_test(self):
        reflection_date=self.ws.cell(row=7, column=3).value
        #print(reflection_date, type(reflection_date))

    def reading_data_line(self, row):
        category_code = str(self.ws.cell(row=row, column=1).value).strip()
        brand_code = str(self.ws.cell(row=row, column=2).value).strip()
        reflection_date = str(self.ws.cell(row=row, column=3).value).strip()
        project_name = str(self.ws.cell(row=row, column=4).value).strip()

        return {
            'category_code': category_code,
            'brand_code': brand_code,
            'reflection_date': reflection_date,
            'project_name': project_name
        }

    def insert_output_file(self, response_data, row):
        """{
                    'new_project_id': new_project_id,
                    'exist_project_name': '',
                    'exist_project_id': '',
                    'exist_reflection_date': '',
                    'project_registration_date': '',
                    'user': '',
                    'error': '',
                    'finishing_time': finishing_time
                }"""
        self.ws.cell(row=row, column=5).value = response_data['new_project_id']
        self.ws.cell(row=row, column=6).value = response_data['exist_project_name']
        self.ws.cell(row=row, column=7).value = response_data['exist_project_id']
        self.ws.cell(row=row, column=8).value = response_data['exist_reflection_date']
        self.ws.cell(row=row, column=9).value = response_data['project_registration_date']
        self.ws.cell(row=row, column=10).value = response_data['user']
        self.ws.cell(row=row, column=11).value = response_data['error']
        self.ws.cell(row=row, column=12).value = response_data['finishing_time']
        self.wb.save(filename=self.output_name)

    def close_excel(self):
        self.wb.close()


class Automation():
    def __init__(self, url, chromedriver_mode):
        self.url = url

        chrome_options = Options()
        chrome_options.add_argument("--disable-notifications")  # 禁用通知
        chrome_options.add_argument("--disable-popup-blocking")  # 禁用弹出窗口阻止
        chrome_options.add_argument("--no-sandbox")  # 避免沙盒模式
        chrome_options.add_argument("--disable-dev-shm-usage")  # 禁用 /dev/shm 使用
        chrome_options.add_argument("--disable-gpu")  # 禁用 GPU 加速
        chrome_options.add_argument("--disable-infobars")  # 禁用信息栏
        chrome_options.add_argument("--disable-extensions")  # 禁用扩展
        chrome_options.add_argument("--disable-browser-side-navigation")  # 禁用浏览器侧边导航
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # 禁用自动控制特性
        #chrome_options.add_argument("--headless") #Hide chrome browser

        if chromedriver_mode == 1:  #"auto"
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                                  options=chrome_options)
        else:    #"manual"
            self.driver = webdriver.Chrome(executable_path='./chromedriver.exe', options=chrome_options)

        self.driver.maximize_window()
        self.wait3 = WebDriverWait(driver=self.driver, timeout=3)
        self.wait5 = WebDriverWait(driver=self.driver, timeout=5)
        self.wait7 = WebDriverWait(driver=self.driver, timeout=7)
        self.wait10 = WebDriverWait(driver=self.driver, timeout=10)
        self.wait20 = WebDriverWait(driver=self.driver, timeout=20)

    def close_browser(self):
        self.driver.close()

    def login(self, ID, PW, subsidiary, ui_language, country):
        print('Login the account...')
        self.driver.get(url=self.url)
        self.driver.delete_all_cookies()


        user_id = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//input[@name="userId"]')))
        user_id.send_keys(ID)
        password = self.wait10.until(
            EC.presence_of_element_located((By.XPATH, '//input[@name="password"]')))
        password.send_keys(PW)
        subsidiary_select = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//select[@name="subsidiaryCd"]')))
        subsidiary_exec = Select(subsidiary_select)
        subsidiary_exec.select_by_value(subsidiary)
        ui_language_select = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//select[@name="uiLanguageCode"]')))
        ui_language_exec = Select(ui_language_select)
        ui_language_exec.select_by_visible_text(str(ui_language).strip().upper())
        login_btn = self.wait5.until(
           EC.presence_of_element_located((By.XPATH, "//button[@onclick=\"$('#form').submit();\"]")))
        login_btn.click()

        ec_btn = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//button[@value="../com/SYS004Page!init.action?systemStr=ec"]')))
        ec_btn.click()

        country_btn = self.wait10.until(EC.presence_of_element_located((By.XPATH, f'//input[@value="{country}"]')))
        country_btn.click()

        print('Login successfully!')

    def main_job(self, category_code, brand_code, reflection_date, project_name):
        """
        category_code = data_line_dict['category_code']
        brand_code = data_line_dict['brand_code']
        reflection_date = data_line_dict['reflection_date']
        project_name = data_line_dict['project_name']
        """
        def step_6(reflection_date, project_name, radio_selector):
            radio_selector.click()

            createProject_btn = self.wait10.until(
                EC.presence_of_element_located((By.XPATH, '//input[@id="new"]'))
            )

            createProject_btn.click()

            calendarInput = self.wait10.until(
                EC.presence_of_element_located((By.XPATH, '//input[@id="konkaiDt"]')))

            calendarInput.clear()
            calendarInput.send_keys(reflection_date)

            projectNameInput = self.wait10.until(
                EC.presence_of_element_located((By.XPATH, '//input[@id="ankenName"]')))

            projectNameInput.clear()
            projectNameInput.send_keys(project_name)

            createJobBtn = self.wait10.until(
                EC.presence_of_element_located((By.XPATH, '//input[@id="createAnken"]')))

            createJobBtn.click()

            try:
                message_area_condition = (
                    EC.presence_of_element_located((By.XPATH, '//div[@id="message_area"]')),
                    EC.text_to_be_present_in_element((By.XPATH, '//div[@id="message_area"]'), 'Please choose a date following')
                )
                message_area = self.wait5.until(lambda driver: all(condition(driver) for condition in message_area_condition))
            except TimeoutException:
                project_id_input = self.wait10.until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="ankenId"]'))
                )
                new_project_id = project_id_input.get_attribute("value")

                current_time = datetime.datetime.now()
                finishing_time = current_time.strftime("%Y/%m/%d %H:%M:%S")

                back_btn = self.wait5.until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="linkb"]'))
                )
                back_btn.click()

                return {
                    'new_project_id': new_project_id,
                    'exist_project_name': '',
                    'exist_project_id': '',
                    'exist_reflection_date': '',
                    'project_registration_date': '',
                    'user': '',
                    'error': '',
                    'finishing_time': finishing_time
                }
            else:

                back_btn = self.wait5.until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="linkb"]'))
                )
                back_btn.click()
                return {
                    'new_project_id': '',
                    'exist_project_name': '',
                    'exist_project_id': '',
                    'exist_reflection_date': '',
                    'project_registration_date': '',
                    'user': '',
                    'error': 'Please input a date following today for Reflection Date.',
                    'finishing_time': ''
                }

        """searchCategoryBrand = self.wait10.until(
            EC.presence_of_element_located((By.XPATH, '//button[@id="new"]')))

        searchCategoryBrand.click()"""
        brand_search_btn = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//button[@value="../ecm/ECM033Page!init.action?returnFlag=0"]')))
        brand_search_btn.click()
        input_categoryCode = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//input[@id="categoryCode"]')))
        input_categoryCode.clear()
        input_categoryCode.send_keys(category_code)

        input_brandCode = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//input[@id="brandCode"]')))
        input_brandCode.clear()
        input_brandCode.send_keys(brand_code)

        radio_reg2 = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//input[@id="regFlg2"]')))
        radio_reg2.click()

        search_btn = self.wait5.until(EC.presence_of_element_located((By.XPATH, '//input[@id="searchBtn"]')))
        search_btn.click()

        try:
            radio_selector = self.wait5.until(EC.presence_of_element_located((By.XPATH, '//table[@class="htCore"]//input[@name="radio"]')))
        except TimeoutException:
            radio_reg1 = self.wait10.until(EC.presence_of_element_located((By.XPATH, '//input[@id="regFlg1"]')))
            radio_reg1.click()

            search_btn.click()
            try:
                radio_selector = self.wait5.until(EC.presence_of_element_located((By.XPATH, '//table[@class="htCore"]//input[@name="radio"]')))
            except TimeoutException:
                back_btn = self.wait5.until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="linkb"]'))
                )
                back_btn.click()
                return {
                    'new_project_id': '',
                    'exist_project_name': '',
                    'exist_project_id': '',
                    'exist_reflection_date': '',
                    'project_registration_date': '',
                    'user': '',
                    'error': 'No corresponding master',
                    'finishing_time': ''
                }

            else:
                return step_6(reflection_date=reflection_date,
                              project_name=project_name,
                              radio_selector=radio_selector)

        else:
            try:
                projectNum_btn = self.wait3.until(EC.presence_of_element_located((By.XPATH, '//table[@class="htCore"]//tbody//a')))
            except TimeoutException:
                return step_6(reflection_date=reflection_date,
                              project_name=project_name,
                              radio_selector=radio_selector)
            else:
                projectNum_btn.click()
                exist_project_name_div = self.wait10.until(
                    EC.presence_of_element_located((By.XPATH, '//table[@id="ankenInfoHead"]//tbody//div[@style="word-wrap:break-word;word-break: break-all;"]'))
                )
                exist_project_name = str(exist_project_name_div.text).strip()

                exist_project_id_input = self.wait10.until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="ankenId"]'))
                )
                exist_project_id = exist_project_id_input.get_attribute('value')

                exist_reflection_date_td = self.wait10.until(
                    EC.presence_of_element_located((By.XPATH, '//td[@id="mukuhyouDt"]'))
                )
                exist_reflection_date = str(exist_reflection_date_td.text).strip()

                project_registration_date_td = self.wait10.until(
                    EC.presence_of_element_located((By.XPATH, "//th[text()='Project Registration Date']/following-sibling::td"))
                )
                project_registration_date = str(project_registration_date_td.text).strip()

                user_td = self.wait10.until(
                    EC.presence_of_element_located((By.XPATH, "//th[text()='User Name']/following-sibling::td"))
                )
                user = str(user_td.text).strip()

                back_btn = self.wait5.until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="linkb"]'))
                )
                back_btn.click()
                return {
                    'new_project_id': '',
                    'exist_project_name': exist_project_name,
                    'exist_project_id': exist_project_id,
                    'exist_reflection_date': exist_reflection_date,
                    'project_registration_date': project_registration_date,
                    'user': user,
                    'error': '',
                    'finishing_time': ''
                }


if __name__ == '__main__':
    # github test
    EXCEL = Excel_Con(file_name='import_sheet.xlsx')
    specs = EXCEL.reading_specs()

    AUTO = Automation(url='https://maint-mdm.jp.misumi-ec.com/mdm/com/COMA01Page!init.action#', chromedriver_mode=0)
    AUTO.login(ID=specs['ID'],
               PW=specs['PW'],
               subsidiary=specs['subsidiary'],
               ui_language=specs['ui_language'],
               country=specs['country'])

    max_row = EXCEL.check_max_row()
    for i in range(10, max_row+1):
        data_line_dict = EXCEL.reading_data_line(row=i)
        category_code = data_line_dict['category_code']
        brand_code = data_line_dict['brand_code']
        reflection_date = data_line_dict['reflection_date']
        project_name = data_line_dict['project_name']

        print(f'\nProcessing data {i-9}/{max_row-9}...\nCategory Code: {category_code}\nBrand Code: {brand_code}\nReflection Date: {reflection_date}\nProject Name: {project_name}\n')
        response_data = AUTO.main_job(category_code=category_code, brand_code=brand_code, reflection_date=reflection_date, project_name=project_name)
        EXCEL.insert_output_file(response_data=response_data, row=i)
        print(f'{response_data}')

    EXCEL.close_excel()

    AUTO.close_browser()
    print('Finished!')
    os.system('pause')